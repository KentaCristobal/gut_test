import openpyxl,xlwings
from datetime import datetime

raw_material='DMTM-P-R-43 RM 201909  素材受払台帳.xlsx'
file_label='Work_order.xlsx'
out_file_label='out_Work_order.xlsx'
file_etiqueta ='RM Etiqueta .xlsx'
out_file_etiqueta='out_RM Etiqueta  .xlsx'


wb_m= openpyxl.load_workbook(raw_material, data_only = True)
ws_m=wb_m['RM_Database MX']
ws_m_return=wb_m.active

wb_ib = openpyxl.load_workbook(file_label)
ws_ib=wb_ib.active


wb_et=openpyxl.load_workbook(file_etiqueta)
ws_et=wb_et.active



cdate = datetime.now().strftime("%Y/%m/%d")
first_response =int(input('Selecciona actividad: 1 Orden de produccion , 2 Etiqueta interna ->>  '))
if first_response ==1:
    mmm=xlwings.Book(raw_material)
    print(mmm)
    ws_ib["Q7"].value = cdate
    ws_ib["E36"].value = cdate

    celda_inico=input('cell desde :  ')
    celda_final=input('cell a :  ')

    numero_orden=input('Registra numero de orden MP～:  ')
    # xlwings.close('DMTM-P-R-43 Entrada y salida de materia prima 201907  素材受払台帳 -.xlsx')
    start=input('Favor de cerrar el archivo y presiona Enter:')


    ws_ib["Q9"].value =numero_orden
    file_custmer= ws_m['G'+ celda_inico +':'+'G'+ celda_final]
    file_dowa= ws_m['H'+ celda_inico +':'+'H'+ celda_final]
    file_part= ws_m['I'+ celda_inico +':'+'I'+ celda_final]
    file_cant= ws_m['R'+ celda_inico +':'+'R'+ celda_final]
    file_tickness= ws_m['N'+ celda_inico +':'+'N'+ celda_final]
    file_width= ws_m['O'+ celda_inico +':'+'O'+ celda_final]
    file_type= ws_m['M'+ celda_inico +':'+'M'+ celda_final]


    for x in range(int(celda_inico)-1,int(celda_final)):
        cdate = datetime.now().strftime("%Y/%m/%d")
        ws_m_return["AP"][x].value=cdate#列だけでなく行も指定す

    for t in range(int(celda_inico)-1,int(celda_final)):
        ws_m_return["AO"][t].value=numero_orden#列だけでなく行も指定す




    for y, row in enumerate(file_dowa):
     for x, cell in enumerate(row):
         if (cell is None) or (cell.value is None): continue
         v = cell.value
         ws_ib.cell(row=16+y+1, column=2+x, value=v)

    for y, row in enumerate(file_part):
     for x, cell in enumerate(row):
         if (cell is None) or (cell.value is None): continue
         v = cell.value
         ws_ib.cell(row=16+y+1, column=6+x, value=v)

    for y, row in enumerate(file_cant):
     for x, cell in enumerate(row):
         if (cell is None) or (cell.value is None): continue
         v = cell.value
         ws_ib.cell(row=16+y+1, column=20+x, value=v)

    for y, row in enumerate(file_tickness):
     for x, cell in enumerate(row):
         if (cell is None) or (cell.value is None): continue
         v = cell.value
         ws_ib.cell(row=16+y+1, column=10+x, value=v)

    for y, row in enumerate(file_width):
     for x, cell in enumerate(row):
         if (cell is None) or (cell.value is None): continue
         v = cell.value
         ws_ib.cell(row=16+y+1, column=12+x, value=v)

    for y, row in enumerate(file_type):
     for x, cell in enumerate(row):
         if (cell is None) or (cell.value is None): continue
         v = cell.value
         ws_ib.cell(row=16+y+1, column=14+x, value=v)

    for y, row in enumerate(file_custmer):
     for x, cell in enumerate(row):
         if (cell is None) or (cell.value is None): continue
         v = cell.value
         ws_ib.cell(row=16+y+1, column=16+x, value=v)

    wb_ib.save(out_file_label)
    wb_m.save(raw_material)

    print('Ok')
    xlwings.Book(raw_material)


elif first_response == 2:
    xlwings.Book(raw_material)
    celda_inico=input('cell desde :  ')
    celda_final=input('cell a :  ')

    file_custmer= ws_m['G'+ celda_inico +':'+'G'+ celda_final]
    file_dowa= ws_m['H'+ celda_inico +':'+'H'+ celda_final]
    file_part= ws_m['I'+ celda_inico +':'+'I'+ celda_final]
    file_cant= ws_m['R'+ celda_inico +':'+'R'+ celda_final]
    file_time= ws_m['A'+ celda_inico +':'+'A'+ celda_final]

    for y, row in enumerate(file_custmer):
     for x, cell in enumerate(row):
         if (cell is None) or (cell.value is None): continue
         v = cell.value
         ws_et.cell(row=0+y+1, column=4+x, value=v)

    for y, row in enumerate(file_dowa):
     for x, cell in enumerate(row):
         if (cell is None) or (cell.value is None): continue
         v = cell.value
         ws_et.cell(row=0+y+1, column=6+x, value=v)


    for y, row in enumerate(file_part):
     for x, cell in enumerate(row):
         if (cell is None) or (cell.value is None): continue
         v = cell.value
         ws_et.cell(row=0+y+1, column=2+x, value=v)

    for y, row in enumerate(file_cant):
     for x, cell in enumerate(row):
         if (cell is None) or (cell.value is None): continue
         v = cell.value
         ws_et.cell(row=0+y+1, column=8+x, value=v)

    for y, row in enumerate(file_time):
     for x, cell in enumerate(row):
         if (cell is None) or (cell.value is None): continue
         v = cell.value
         ws_et.cell(row=0+y+1, column=1+x, value=v)
    wb_et.save(out_file_etiqueta)
    xlwings.Book(raw_material)
    print('ok')


else:
    print( "incorrect data")
